"""
ChinaAMC: resolve holdings xlsx from product page or direct download URL.

Examples:
  Product page:
    https://www.chinaamc.com.hk/zh-hant/product/chinaamc-hk-us-ai-etf-3140/
  Direct download:
    https://www.chinaamc.com.hk/wp-content/uploads/chinaamc/holdings/HKUSAIF_TC.xlsx?v=1777457002
"""

from __future__ import annotations

import io
import re
from datetime import datetime
from typing import List, Optional, Sequence
from urllib.parse import parse_qs, urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from model.page_result import PageResult
from scrapers.scraper_utils import extract_as_of_from_cells, fetch_content, fetch_text, normalize_row


def _is_download_url(url: str) -> bool:
    u = url.lower()
    return ("/holdings/" in u and ".xlsx" in u) or u.endswith(".xlsx")


def _resolve_download_urls(input_url: str, html: Optional[str] = None) -> List[str]:
    if _is_download_url(input_url):
        return [input_url]

    if html is None:
        return []

    soup = BeautifulSoup(html, "html.parser")
    out: List[str] = []
    for a in soup.find_all("a"):
        href = (a.get("href") or "").strip()
        if not href:
            continue
        h = href.lower()
        if "/holdings/" in h and ".xlsx" in h:
            out.append(urljoin(input_url, href))

    # Keep order, dedupe.
    seen = set()
    deduped: List[str] = []
    for u in out:
        if u not in seen:
            seen.add(u)
            deduped.append(u)
    return deduped


def _etf_code(download_url: str, source_url: str) -> str:
    # Prefer file stem, e.g. HKUSAIF_TC.xlsx -> HKUSAIF
    path = urlparse(download_url).path
    name = path.split("/")[-1]
    m = re.match(r"([A-Za-z0-9]+)(?:_[A-Za-z0-9]+)?\.xlsx$", name, flags=re.I)
    if m:
        return m.group(1)

    q = parse_qs(urlparse(download_url).query)
    for key in ("fundId", "fund_id", "code"):
        value = (q.get(key) or [""])[0].strip()
        if value:
            return value

    parts = [p for p in urlparse(source_url).path.split("/") if p]
    return parts[-1] if parts else ""


def _is_header_row(cells: List[str]) -> bool:
    if not cells:
        return False
    joined = " ".join(cells).lower()
    return ("名稱" in joined or "name" in joined) and ("證券代碼" in joined or "ticker" in joined) and (
        "比重" in joined or "%" in joined or "weight" in joined
    )


def _parse_xlsx(content: bytes) -> tuple[List[List[str]], Optional[str]]:
    wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    try:
        ws = wb.active
        rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = normalize_row(row)
            if any(cells):
                rows.append(cells)
    finally:
        wb.close()

    as_of = extract_as_of_from_cells(rows, max_rows=10)
    if not as_of:
        for row in rows[:10]:
            for cell in row:
                m = re.search(r"(\d{4})/(\d{1,2})/(\d{1,2})", cell)
                if m:
                    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
                    as_of = f"{y:04d}{mo:02d}{d:02d}"
                    break
            if as_of:
                break
    header_idx: Optional[int] = None
    for i, row in enumerate(rows):
        if _is_header_row(row):
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("Cannot find ChinaAMC holdings header row.")

    header = rows[header_idx]
    data_rows = rows[header_idx + 1 :]
    table = [header] + [r for r in data_rows if any(c for c in r)]
    if len(table) <= 1:
        raise ValueError("ChinaAMC holdings file has header but no data rows.")
    return table, as_of


def _download_and_parse(
    session: requests.Session,
    download_url: str,
    timeout: int,
) -> tuple[List[List[str]], Optional[str]]:
    content = fetch_content(
        session,
        download_url,
        timeout,
        accept="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
    )
    return _parse_xlsx(content)


class ChinaamcScraper:
    """Fetch ChinaAMC holdings from product page or direct xlsx URL."""

    def __init__(self, timeout: int = 45, session: Optional[requests.Session] = None):
        self.timeout = timeout
        self.session = session or requests.Session()

    def scrape_url(self, url: str) -> PageResult:
        code: Optional[str] = None
        try:
            main_html = None if _is_download_url(url) else fetch_text(self.session, url, self.timeout)
            download_urls = _resolve_download_urls(url, main_html)
            if not download_urls:
                raise ValueError("Cannot find ChinaAMC holdings download URL on product page.")

            # Use first resolved download link.
            download_url = download_urls[0]
            code = _etf_code(download_url, url)
            rows, as_of = _download_and_parse(self.session, download_url, self.timeout)
            if not as_of:
                as_of = datetime.now().strftime("%Y%m%d")

            return PageResult(
                url=url,
                ok=True,
                etf_code=code,
                as_of_date=as_of,
                rows=rows,
            )
        except Exception as e:
            return PageResult(url=url, ok=False, etf_code=code, error=str(e))

    def scrape_urls(self, urls: Sequence[str]) -> List[PageResult]:
        return [self.scrape_url(u) for u in urls]

