"""
Bosera: download holdings from direct file URLs (XLSX or CSV).

Example XLSX:
  https://www.bosera.com.hk/api/fundinfo/exportholdingsexcel.do?language=hk&fundCode=USDMMETF
"""

from __future__ import annotations

import csv
import io
import re
from typing import List, Optional, Sequence
from urllib.parse import parse_qs, unquote, urlparse

import requests
from openpyxl import load_workbook

from model.page_result import PageResult

_DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,text/csv,*/*",
}

def _etf_code_from_bosera_path(url: str) -> Optional[str]:
    query = unquote(urlparse(url).query)
    return parse_qs(query).get("fundCode", [None])[0]

def _as_of_yyyymmdd_from_zh(text: str) -> Optional[str]:
    m = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", text)
    if not m:
        return None
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    return f"{y:04d}{mo:02d}{d:02d}"

def _as_of_from_bosera_cells(rows: List[List[str]]) -> Optional[str]:
    """DD/MM/YYYY in title rows (e.g. As at 14/04/2026)."""
    for row in rows[:20]:
        for cell in row:
            if not cell:
                continue
            s = str(cell)
            m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
            if m:
                d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
                return f"{y:04d}{mo:02d}{d:02d}"
            d2 = _as_of_yyyymmdd_from_zh(s)
            if d2:
                return d2
    return None

def _cell_str(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _normalize_row(row: Sequence[object]) -> List[str]:
    out: List[str] = []
    for c in row:
        s = _cell_str(c)
        s = s.replace("\n", " ").strip()
        out.append(s)
    return out


def _is_holdings_header_row(cells: List[str]) -> bool:
    if not cells or not cells[0]:
        return False
    first = cells[0].lower()
    return (
        "net assets" in first
        or "資產淨值" in cells[0]
    )


def _sheet_to_row_grid(ws) -> List[List[str]]:
    rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = _normalize_row(row)
        if any(c for c in cells):
            rows.append(cells)
    return rows


def _parse_holdings_table(rows: List[List[str]]) -> tuple[List[List[str]], Optional[str]]:
    as_of = _as_of_from_bosera_cells(rows)

    header_idx: Optional[int] = None
    for i, row in enumerate(rows):
        if _is_holdings_header_row(row):
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("Cannot find holdings header row (Net Assets Holding / 資產淨值).")

    header = rows[header_idx]
    data_rows = rows[header_idx + 1 :]
    table = [header] + [r for r in data_rows if any(c for c in r)]

    if len(table) <= 1:
        raise ValueError("Table has header but no data rows.")

    return table, as_of


def _parse_holdings_csv(text: str) -> tuple[List[List[str]], Optional[str]]:
    buf = io.StringIO(text.lstrip("\ufeff"))
    reader = csv.reader(buf)
    raw_rows: List[List[str]] = [
        [_cell_str(c).replace("\n", " ") for c in row]
        for row in reader
        if any((c or "").strip() for c in row)
    ]

    as_of: Optional[str] = None
    for row in raw_rows[:8]:
        for cell in row:
            d = _as_of_yyyymmdd_from_zh(cell)
            if d:
                as_of = d
                break
            m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", cell)
            if m:
                d0, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
                as_of = f"{y:04d}{mo:02d}{d0:02d}"
                break
        if as_of:
            break

    header_idx: Optional[int] = None
    for i, row in enumerate(raw_rows):
        if _is_holdings_header_row([(c or "").strip() for c in row]):
            header_idx = i
            break
        first = (row[0] or "").strip().lower()
        if first == "ticker":
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("Cannot find holdings header row in CSV.")

    header = [c.strip() for c in raw_rows[header_idx]]
    data_rows = raw_rows[header_idx + 1 :]
    table = [header] + [[c.strip() for c in r] for r in data_rows if any((c or "").strip() for c in r)]

    if len(table) <= 1:
        raise ValueError("CSV has header but no data rows.")

    return table, as_of


def _is_xlsx_content(content: bytes, url: str) -> bool:
    if url.lower().split("?")[0].endswith(".xlsx"):
        return True
    return len(content) >= 4 and content[:2] == b"PK"

class BoseraScraper:
    """Placeholder scraper for Bosera pages."""

    def __init__(self, timeout: int = 60, session: Optional[requests.Session] = None):
        self.timeout = timeout
        self.session = session or requests.Session()

    def scrape_url(self, url: str) -> PageResult:
        etf_code = _etf_code_from_bosera_path(url)
        try:
            resp = self.session.get(url, timeout=self.timeout, headers=_DEFAULT_HEADERS)
            resp.raise_for_status()
            content = resp.content

            if _is_xlsx_content(content, url):
                bio = io.BytesIO(content)
                # read_only mode openpyxl may expose only A1. Use normal mode.
                wb = load_workbook(bio, read_only=False, data_only=True)
                try:
                    ws = wb.active
                    grid = _sheet_to_row_grid(ws)
                finally:
                    wb.close()
                rows, as_of = _parse_holdings_table(grid)
            else:
                text = content.decode("utf-8-sig", errors="replace")
                rows, as_of = _parse_holdings_csv(text)

            if not etf_code and rows and len(rows) > 1:
                etf_code = rows[1][0].strip('"') if rows[1] else None

            return PageResult(
                url=url,
                ok=True,
                etf_code=etf_code,
                as_of_date=as_of,
                rows=rows,
            )
        except Exception as e:
            return PageResult(url=url, ok=False, etf_code=etf_code, error=str(e))

    def scrape_urls(self, urls: Sequence[str]) -> List[PageResult]:
        return [self.scrape_url(u) for u in urls]
