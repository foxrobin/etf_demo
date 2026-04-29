"""
Premia: product page exposes holdings Excel download via exportHoldings().

Example:
  https://etfprod.premia-partners.com/etf/tc/2803
  -> /etf/exportHoldingsToExcel?fundId=CXBS&lang=hk
"""

from __future__ import annotations

import io
import re
from datetime import datetime
from typing import List, Optional, Sequence
from urllib.parse import parse_qs, urljoin, urlparse

import requests
import xlrd
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from model.page_result import PageResult
from scrapers.scraper_utils import (
    cell_str,
    extract_as_of_from_cells,
    extract_as_of_yyyymmdd,
    fetch_content,
    fetch_text,
    is_xlsx_content,
    normalize_row,
)

_DOWNLOAD_ACCEPT = "application/vnd.ms-excel,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*"
_DIRECT_DOWNLOAD_HINT = "/etf/exportHoldingsToExcel"
_DOWNLOAD_BASE_URL = "https://etfprod.premia-partners.com/etf/exportHoldingsToExcel"

# Fallback mapping for pages where download URL is not rendered in HTML.
_FUND_ID_BY_PRODUCT_CODE = {
    "2803": "CXBS",
}


def _is_download_url(url: str) -> bool:
    return _DIRECT_DOWNLOAD_HINT.lower() in url.lower()


def _lang_from_page_url(page_url: str) -> str:
    # /etf/tc/2803 -> hk ; /etf/en/2803 -> en ; /etf/sc/2803 -> cn
    parts = [p for p in urlparse(page_url).path.split("/") if p]
    locale = parts[1].lower() if len(parts) >= 2 else "tc"
    if locale == "en":
        return "en"
    if locale in {"sc", "cn"}:
        return "cn"
    return "hk"


def _build_download_url(fund_id: str, page_url: str) -> str:
    lang = _lang_from_page_url(page_url)
    return f"{_DOWNLOAD_BASE_URL}?fundId={fund_id}&lang={lang}"


def _fallback_download_url_by_product_code(page_url: str) -> Optional[str]:
    parts = [p for p in urlparse(page_url).path.split("/") if p]
    code = parts[-1] if parts else ""
    fund_id = _FUND_ID_BY_PRODUCT_CODE.get(code)
    if not fund_id:
        return None
    return _build_download_url(fund_id, page_url)


def _extract_download_url(main_html: str, page_url: str) -> str:
    soup = BeautifulSoup(main_html, "html.parser")

    # Sometimes URL already appears in anchors.
    for a in soup.find_all("a"):
        href = (a.get("href") or "").strip()
        if href and _is_download_url(href):
            return urljoin(page_url, href)

    # Sometimes URL appears in scripts.
    m = re.search(r"/etf/exportHoldingsToExcel\?fundId=([A-Za-z0-9_-]+)(?:&lang=([A-Za-z]+))?", main_html)
    if m:
        fund_id = m.group(1)
        return _build_download_url(fund_id, page_url)

    fallback_url = _fallback_download_url_by_product_code(page_url)
    if fallback_url:
        return fallback_url

    raise ValueError("Cannot find Premia holdings download URL on product page.")


def _etf_code(download_url: str, source_url: str) -> str:
    q = parse_qs(urlparse(download_url).query)
    fund_id = (q.get("fundId") or [""])[0].strip()
    if fund_id:
        return fund_id

    parts = [p for p in urlparse(source_url).path.split("/") if p]
    return parts[-1] if parts else ""


def _is_header_row(cells: List[str]) -> bool:
    if not cells:
        return False
    joined = " ".join(cells).lower()
    has_code = ("證券代碼" in joined) or ("security code" in joined) or ("ticker" in joined)
    has_name = ("成份股名稱" in joined) or ("名稱" in joined) or ("name" in joined)
    has_weight = ("資產淨值百分比" in joined) or ("持倉比重" in joined) or ("weight" in joined) or ("%" in joined)
    return (has_code and has_name) or (has_name and has_weight)


def _parse_rows_grid(rows: List[List[str]]) -> tuple[List[List[str]], Optional[str]]:
    as_of = extract_as_of_from_cells(rows, max_rows=20)

    header_idx: Optional[int] = None
    for i, row in enumerate(rows):
        if _is_header_row(row):
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("Cannot find Premia holdings header row in downloaded file.")

    header = rows[header_idx]
    data_rows = rows[header_idx + 1 :]
    table = [header] + [r for r in data_rows if any(c for c in r)]
    if len(table) <= 1:
        raise ValueError("Holdings file has header but no data rows.")
    return table, as_of


def _parse_xlsx(content: bytes) -> tuple[List[List[str]], Optional[str]]:
    wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    try:
        ws = wb.active
        rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = normalize_row(row)
            if any(cells):
                rows.append(cells)
    finally:
        wb.close()
    return _parse_rows_grid(rows)


def _parse_xls(content: bytes) -> tuple[List[List[str]], Optional[str]]:
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)
    rows: List[List[str]] = []
    for r in range(ws.nrows):
        vals = [cell_str(ws.cell_value(r, c)) for c in range(ws.ncols)]
        row = [v.replace("\xa0", " ").replace("\n", " ").strip() for v in vals]
        if any(row):
            rows.append(row)
    return _parse_rows_grid(rows)


def _parse_html_table(content: bytes) -> tuple[List[List[str]], Optional[str]]:
    text = content.decode("utf-8-sig", errors="replace")
    soup = BeautifulSoup(text, "html.parser")
    best_rows: List[List[str]] = []
    for table in soup.find_all("table"):
        headers = [cell_str(th.get_text()) for th in table.find_all("th")]
        if not _is_header_row(headers):
            continue
        rows: List[List[str]] = []
        if headers:
            rows.append(headers)
        tr_list = table.find("tbody").find_all("tr") if table.find("tbody") else table.find_all("tr")
        for tr in tr_list:
            row = [cell_str(c.get_text()) for c in tr.find_all(["td", "th"])]
            if any(row):
                rows.append(row)
        if len(rows) > len(best_rows):
            best_rows = rows

    if len(best_rows) <= 1:
        raise ValueError("Downloaded Premia file is neither parseable XLS/XLSX nor HTML table.")
    as_of = extract_as_of_yyyymmdd(text)
    return best_rows, as_of


def _parse_download(content: bytes, download_url: str) -> tuple[List[List[str]], Optional[str]]:
    if is_xlsx_content(content, download_url):
        return _parse_xlsx(content)
    try:
        return _parse_xls(content)
    except Exception:
        return _parse_html_table(content)


class PremiaScraper:
    """Fetch Premia holdings by resolving and downloading holdings Excel file."""

    def __init__(self, timeout: int = 45, session: Optional[requests.Session] = None):
        self.timeout = timeout
        self.session = session or requests.Session()

    def scrape_url(self, url: str) -> PageResult:
        download_url = ""
        code: Optional[str] = None
        try:
            if _is_download_url(url):
                download_url = url
            else:
                main_html = fetch_text(self.session, url, self.timeout)
                download_url = _extract_download_url(main_html, url)

            code = _etf_code(download_url, url)
            content = fetch_content(self.session, download_url, self.timeout, accept=_DOWNLOAD_ACCEPT)
            rows, as_of = _parse_download(content, download_url)
            if not as_of:
                as_of = datetime.now().strftime("%Y%m%d")

            return PageResult(
                url=url,
                ok=True,
                etf_code=code,
                as_of_date=as_of,
                rows=rows,
            )
        except Exception as e:
            return PageResult(url=url, ok=False, etf_code=code, error=str(e))

    def scrape_urls(self, urls: Sequence[str]) -> List[PageResult]:
        return [self.scrape_url(u) for u in urls]
