"""
CSOP: product page links to holdings download (.xls/.xlsx) via anchor title.

Example:
  https://www.csopasset.com/tc/products/china-a50-etf
  -> title="https://website-api.csopasset.com/cmsApi/Holdings/product/list/download?fundId=CO-A50F"
"""

from __future__ import annotations

import io
import re
from datetime import datetime
from typing import List, Optional, Sequence
from urllib.parse import parse_qs, urljoin, urlparse

import requests
import xlrd
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from model.page_result import PageResult
from scrapers.scraper_utils import (
    cell_str,
    extract_as_of_from_cells,
    extract_as_of_yyyymmdd,
    fetch_content,
    fetch_text,
    is_xlsx_content,
    normalize_row,
)

_DOWNLOAD_ACCEPT = "application/vnd.ms-excel,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*"
_DIRECT_DOWNLOAD_HINT = "/cmsApi/Holdings/product/list/download"
_DOWNLOAD_BASE_URL = "https://website-api.csopasset.com/cmsApi/Holdings/product/list/download"

# Product slug fallback for cases where product page blocks scraping and no anchor is visible.
_FUND_ID_BY_SLUG = {
    "china-a50-etf": "CO-A50F",
    "hk-nik225":"HK-NIK225"
}


def _clean_text(text: str) -> str:
    return " ".join(text.replace("\xa0", " ").split())


def _is_csop_download_url(url: str) -> bool:
    return _DIRECT_DOWNLOAD_HINT.lower() in url.lower()


def _build_download_url(fund_id: str) -> str:
    return f"{_DOWNLOAD_BASE_URL}?fundId={fund_id}"


def _fallback_download_url_by_slug(page_url: str) -> Optional[str]:
    parts = [p for p in urlparse(page_url).path.split("/") if p]
    slug = parts[-1].strip().lower() if parts else ""
    fund_id = _FUND_ID_BY_SLUG.get(slug)
    return _build_download_url(fund_id) if fund_id else None


def _extract_download_url(main_html: str, page_url: str) -> str:
    soup = BeautifulSoup(main_html, "html.parser")

    for a in soup.find_all("a"):
        cls = " ".join(a.get("class") or []).lower()
        title = (a.get("title") or "").strip()
        href = (a.get("href") or "").strip()

        if "downloadholdings" in cls and title and _is_csop_download_url(title):
            return title
        if href and _is_csop_download_url(href):
            return urljoin(page_url, href)

    m = re.search(r"https://website-api\.csopasset\.com/cmsApi/Holdings/product/list/download\?[^\s\"'>]+", main_html)
    if m:
        return m.group(0)

    slug_fallback = _fallback_download_url_by_slug(page_url)
    if slug_fallback:
        return slug_fallback

    raise ValueError("Cannot find CSOP holdings download URL on product page.")


def _etf_code_from_urls(download_url: str, source_url: str) -> str:
    q = parse_qs(urlparse(download_url).query)
    fund_id = (q.get("fundId") or [""])[0].strip()
    if fund_id:
        return fund_id

    base_parts = [p for p in urlparse(source_url).path.split("/") if p]
    return base_parts[-1] if base_parts else ""


def _is_holdings_header_row(cells: List[str]) -> bool:
    if not cells:
        return False
    joined = " ".join(cells).lower()
    has_weight = ("weight" in joined) or ("權重" in joined)
    has_name_or_code = ("stock" in joined) or ("name" in joined) or ("代號" in joined) or ("名稱" in joined)
    return has_weight and has_name_or_code


def _parse_table_rows(rows: List[List[str]]) -> tuple[List[List[str]], Optional[str]]:
    as_of = extract_as_of_from_cells(rows, max_rows=20)

    header_idx: Optional[int] = None
    for i, row in enumerate(rows):
        if _is_holdings_header_row(row):
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("Cannot find CSOP holdings header row in downloaded file.")

    header = rows[header_idx]
    data_rows = rows[header_idx + 1 :]
    table = [header] + [r for r in data_rows if any(c for c in r)]
    if len(table) <= 1:
        raise ValueError("Holdings file has header but no data rows.")

    return table, as_of


def _parse_from_xlsx(content: bytes) -> tuple[List[List[str]], Optional[str]]:
    wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    try:
        ws = wb.active
        rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = normalize_row(row)
            if any(cells):
                rows.append(cells)
    finally:
        wb.close()
    return _parse_table_rows(rows)


def _parse_from_xls(content: bytes) -> tuple[List[List[str]], Optional[str]]:
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)
    rows: List[List[str]] = []
    for r in range(ws.nrows):
        vals = [cell_str(ws.cell_value(r, c)) for c in range(ws.ncols)]
        norm = [_clean_text(v) for v in vals]
        if any(norm):
            rows.append(norm)
    return _parse_table_rows(rows)


def _parse_from_html_table(content: bytes) -> tuple[List[List[str]], Optional[str]]:
    text = content.decode("utf-8-sig", errors="replace")
    soup = BeautifulSoup(text, "html.parser")
    best_rows: List[List[str]] = []
    for table in soup.find_all("table"):
        headers = [_clean_text(th.get_text()) for th in table.find_all("th")]
        if not _is_holdings_header_row(headers):
            continue

        rows: List[List[str]] = []
        if headers:
            rows.append(headers)
        tr_list = table.find("tbody").find_all("tr") if table.find("tbody") else table.find_all("tr")
        for tr in tr_list:
            row = [_clean_text(c.get_text()) for c in tr.find_all(["td", "th"])]
            if any(row):
                rows.append(row)
        if len(rows) > len(best_rows):
            best_rows = rows

    if len(best_rows) <= 1:
        raise ValueError("Downloaded CSOP file is neither parseable XLS/XLSX nor HTML table.")

    as_of = extract_as_of_yyyymmdd(text)
    return best_rows, as_of


def _parse_downloaded_file(content: bytes, download_url: str) -> tuple[List[List[str]], Optional[str]]:
    if is_xlsx_content(content, download_url):
        return _parse_from_xlsx(content)

    try:
        return _parse_from_xls(content)
    except Exception:
        return _parse_from_html_table(content)


class CsopScraper:
    """Fetch CSOP holdings by resolving and downloading holdings file."""

    def __init__(self, timeout: int = 45, session: Optional[requests.Session] = None):
        self.timeout = timeout
        self.session = session or requests.Session()

    def scrape_url(self, url: str) -> PageResult:
        download_url = ""
        etf_code: Optional[str] = None
        try:
            if _is_csop_download_url(url):
                download_url = url
            else:
                main_html = fetch_text(self.session, url, self.timeout)
                download_url = _extract_download_url(main_html, url)

            etf_code = _etf_code_from_urls(download_url, url)
            content = fetch_content(self.session, download_url, self.timeout, accept=_DOWNLOAD_ACCEPT)
            rows, as_of = _parse_downloaded_file(content, download_url)
            if not as_of:
                as_of = datetime.now().strftime("%Y%m%d")

            return PageResult(
                url=url,
                ok=True,
                etf_code=etf_code,
                as_of_date=as_of,
                rows=rows,
            )
        except Exception as e:
            return PageResult(url=url, ok=False, etf_code=etf_code, error=str(e))

    def scrape_urls(self, urls: Sequence[str]) -> List[PageResult]:
        return [self.scrape_url(u) for u in urls]
